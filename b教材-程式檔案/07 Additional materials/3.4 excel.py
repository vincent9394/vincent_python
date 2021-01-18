# openpyxl 參考文檔
# https://openpyxl.readthedocs.io/en/stable/

# 匯入 openpyxl 模組
from openpyxl import *

# 取得 Workbook
wb = Workbook()                       # 開新檔案
wb = load_workbook('result.xlsx')   	# 讀取現有檔案

# 取得 Worksheet
ws = wb.active						     # 讀取第一張 Worksheet
ws = wb['Sheet']					     # 直接開啟某張 Worksheet
ws = wb.create_sheet('Sheet2')	  # 開新 Worksheet Sheet2

# 删除 Worksheet
del wb['Sheet2']					# 删除 Sheet2

# 取得 Worksheet
ws = wb['Sheet']					# 讀取Worksheet Sheet

# 新增資料行
ws.append( ['id', 'name'] )
ws.append( [1, 'benson'] )
ws.append( [2, 'kong'] )
ws.append( [3, 'tom'] )

# 取得表格行列數目
print( ws.max_row )			# 返回 4
print( ws.max_column )		# 返回 2
print( ws.min_row )			# 返回 1
print( ws.min_column )		# 返回 1

# 儲存格定位
print( ws['A1'] )			   # 取得A1格
print( ws['A1:B2'] )		   # 取得A1至B2格
print( ws['1:2'] )			# 取得1至2行
print( ws['A:B'] )			# 取得A至B列

# 儲存格定位 by index
print( ws.cell(1, 1) )		# 返回 A1 格
print( ws.cell(3, 2) )		# 返回 B3 格

# 取得儲存格資料
print( ws['B3'].value )				   # 返回 kong
print( ws.cell(3, 2).value )		   # 返回 kong
print( ws['A1:B3'][2][1].value )	# 返回 kong

# 把儲存格 B3 設定為 'mary'
ws['B3'].value = 'mary'
ws.cell(3, 2).value = 'mary'
ws.cell(3, 2, 'mary')

# 把儲存格 A3 設定為算式 A2+1
ws['A3'].value = '=A2+1'

# 删除儲存格 A4 
del ws['A4']

# 儲存檔案
wb.save( 'result.xlsx' )



